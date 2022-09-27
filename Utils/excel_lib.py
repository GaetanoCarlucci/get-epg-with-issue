import os
import openpyxl
from openpyxl import Workbook
import xlrd

class Excel:
    def __init__(self, workbook_path, workbook_name):
        self.column_names = ""
        self.wb = Workbook()
        self.workbook_file = workbook_path + "/" + workbook_name + ".xlsx"
        Workbook.save(self.wb, self.workbook_file)

    def create_sheet(self, sheet_name, column_names):
        xl = openpyxl.load_workbook(self.workbook_file)
        xl.create_sheet(sheet_name, 0)
        xl_sheet = xl[sheet_name]
        column = 1
        for name in column_names:
            xl_sheet.cell(row=1, column=column, value=name)
            column += 1
        xl.save(self.workbook_file)

    def fill_sheet(self, data, sheet_name, row = None):
        #read the header
        xl_xlrd = xlrd.open_workbook(self.workbook_file)
        xl_sheet = xl_xlrd.sheet_by_name(sheet_name=sheet_name)
        xl_file_header = xl_sheet.row(rowx=0)

        # prepare for writing
        xl = openpyxl.load_workbook(self.workbook_file)
        xl_sheet = xl[sheet_name]

        if row == None:
            row = 2
        for xl_row in range(len(data)):
            column = 1
            r = row + xl_row
            for header in xl_file_header:
                xl_sheet.cell(row=r, column=column, value=data[xl_row][header.value])
                column += 1
        xl.save(self.workbook_file)
        return r

    def convert_to_csv(self, sheet_name, csv_file_name):

        # open current sheet to conver in csv
        xl = openpyxl.load_workbook(self.workbook_file)
        xl_sheet = xl[sheet_name]

        csv = open(csv_file_name, "w+")

        for row in xl_sheet:
            l = list(row)
            for i in range(len(l)):
                if i == len(l) - 1:
                    csv.write(str(l[i].value))
                else:
                    csv.write(str(l[i].value) + ',')
            csv.write('\n')

        ## close the csv file
        csv.close()